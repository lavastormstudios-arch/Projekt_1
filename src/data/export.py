import os
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.models.base_entry import Entry
from src.models.enums import EntryType, EntryStatus
from src.utils.date_helpers import format_date
from src.utils.constants import DATA_DIR


def export_entries(
    entries: List[Entry],
    filepath: Optional[str] = None,
    title: str = "WKZ & Bonus Report"
) -> str:
    if filepath is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = os.path.join(DATA_DIR, f"export_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    status_fills = {
        EntryStatus.OFFEN: PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        EntryStatus.ABGERECHNET: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        EntryStatus.UEBERFAELLIG: PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        EntryStatus.STORNIERT: PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    }

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A2"] = f"Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A3"] = f"Anzahl Einträge: {len(entries)}"

    # Headers
    headers = ["ID", "Typ", "Lieferant", "Beschreibung", "Status",
               "Betrag (erw.)", "Betrag (abger.)", "Abrechnungsfrist", "Abgerechnet am", "Notizen"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Data rows
    for row_idx, entry in enumerate(entries, 6):
        values = [
            entry.id, entry.entry_type.value, entry.supplier_name,
            entry.description, entry.status.value,
            entry.amount, entry.amount_billed,
            format_date(entry.billing_deadline),
            format_date(entry.date_billed), entry.notes
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            if col == 5:  # Status column
                fill = status_fills.get(entry.status)
                if fill:
                    cell.fill = fill

    # Column widths
    widths = [10, 14, 20, 30, 14, 14, 14, 16, 16, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=5, column=i).column_letter].width = w

    # Summary sheet
    ws2 = wb.create_sheet("Zusammenfassung")
    ws2["A1"] = "Zusammenfassung"
    ws2["A1"].font = title_font

    summary_data = []
    for entry_type in EntryType:
        type_entries = [e for e in entries if e.entry_type == entry_type]
        if type_entries:
            total = sum(e.amount for e in type_entries)
            billed = sum(e.amount_billed for e in type_entries)
            summary_data.append((entry_type.value, len(type_entries), total, billed))

    ws2.cell(row=3, column=1, value="Typ").font = Font(bold=True)
    ws2.cell(row=3, column=2, value="Anzahl").font = Font(bold=True)
    ws2.cell(row=3, column=3, value="Erwartet").font = Font(bold=True)
    ws2.cell(row=3, column=4, value="Abgerechnet").font = Font(bold=True)

    for i, (typ, count, total, billed) in enumerate(summary_data, 4):
        ws2.cell(row=i, column=1, value=typ)
        ws2.cell(row=i, column=2, value=count)
        ws2.cell(row=i, column=3, value=total)
        ws2.cell(row=i, column=4, value=billed)

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14

    wb.save(filepath)
    return filepath
