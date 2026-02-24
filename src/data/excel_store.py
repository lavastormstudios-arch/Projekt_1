import os
from datetime import date
from typing import List, Optional

from openpyxl import Workbook, load_workbook

from src.models.base_entry import Entry
from src.models.enums import EntryType, EntryStatus
from src.models.supplier import Supplier
from src.utils.constants import (
    DATA_DIR, ENTRIES_FILE, SUPPLIERS_FILE,
    USERS_FILE, DEPARTMENTS, ROLES, MODULES, ACTION_PERMISSIONS,
    ENTRY_COLUMNS, SUPPLIER_COLUMNS
)
from src.utils.date_helpers import parse_date, format_date, safe_float, safe_int

# Default permission tables
_DEFAULT_MODULE_ACCESS = {
    "CM":        {"WKZ & Bonus": True,  "Lieferantenmanagement": True},
    "Marketing": {"WKZ & Bonus": False, "Lieferantenmanagement": True},
    "Vertrieb":  {"WKZ & Bonus": True,  "Lieferantenmanagement": True},
    "SCM":       {"WKZ & Bonus": True,  "Lieferantenmanagement": True},
}

_DEFAULT_ACTION_RIGHTS = {
    "Admin":            {"can_edit": True,  "can_delete": True,  "can_invoice": True,  "can_export": True,  "can_import": True,  "is_admin": True},
    "Abteilungsleiter": {"can_edit": True,  "can_delete": True,  "can_invoice": True,  "can_export": True,  "can_import": True,  "is_admin": False},
    "Teamleiter":       {"can_edit": True,  "can_delete": False, "can_invoice": True,  "can_export": True,  "can_import": False, "is_admin": False},
    "Mitarbeiter":      {"can_edit": False, "can_delete": False, "can_invoice": False, "can_export": True,  "can_import": False, "is_admin": False},
}


class ExcelStore:
    def __init__(self):
        os.makedirs(DATA_DIR, exist_ok=True)
        self._ensure_file(ENTRIES_FILE, ENTRY_COLUMNS)
        self._ensure_file(SUPPLIERS_FILE, SUPPLIER_COLUMNS)

    def _ensure_file(self, path: str, columns: list):
        if not os.path.exists(path):
            wb = Workbook()
            ws = wb.active
            ws.append(columns)
            wb.save(path)

    # --- Entries ---

    def load_entries(self) -> List[Entry]:
        wb = load_workbook(ENTRIES_FILE)
        ws = wb.active
        entries = []
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            data = dict(zip(headers, row))
            entry = Entry(
                id=str(data.get("id", "")),
                entry_type=EntryType(data.get("entry_type", "WKZ")),
                supplier_id=str(data.get("supplier_id", "")),
                supplier_name=str(data.get("supplier_name", "")),
                description=str(data.get("description", "") or ""),
                status=EntryStatus(data.get("status", "Offen")),
                amount=safe_float(data.get("amount")),
                amount_billed=safe_float(data.get("amount_billed")),
                date_start=parse_date(data.get("date_start")),
                date_end=parse_date(data.get("date_end")),
                billing_deadline=parse_date(data.get("billing_deadline")),
                date_billed=parse_date(data.get("date_billed")),
                kickback_articles=str(data.get("kickback_articles", "") or ""),
                umsatzbonus_staffeln=str(data.get("umsatzbonus_staffeln", "") or ""),
                wkz_is_percentage=bool(data.get("wkz_is_percentage", False)),
                wkz_percentage=safe_float(data.get("wkz_percentage")),
                wkz_category=str(data.get("wkz_category", "") or ""),
                notes=str(data.get("notes", "") or ""),
                created_at=str(data.get("created_at", "") or ""),
                invoice_number=str(data.get("invoice_number", "") or ""),
            )
            entries.append(entry)
        wb.close()
        return entries

    def save_entries(self, entries: List[Entry]):
        wb = Workbook()
        ws = wb.active
        ws.append(ENTRY_COLUMNS)
        for e in entries:
            ws.append([
                e.id, e.entry_type.value, e.supplier_id, e.supplier_name,
                e.description, e.status.value, e.amount, e.amount_billed,
                format_date(e.date_start), format_date(e.date_end),
                format_date(e.billing_deadline), format_date(e.date_billed),
                e.kickback_articles,
                e.umsatzbonus_staffeln,
                e.wkz_is_percentage, e.wkz_percentage,
                e.wkz_category, e.notes, e.created_at, e.invoice_number
            ])
        wb.save(ENTRIES_FILE)

    def add_entry(self, entry: Entry):
        entries = self.load_entries()
        entries.append(entry)
        self.save_entries(entries)

    def update_entry(self, entry: Entry):
        entries = self.load_entries()
        for i, e in enumerate(entries):
            if e.id == entry.id:
                entries[i] = entry
                break
        self.save_entries(entries)

    def delete_entry(self, entry_id: str):
        entries = [e for e in self.load_entries() if e.id != entry_id]
        self.save_entries(entries)

    # --- Suppliers ---

    def load_suppliers(self) -> List[Supplier]:
        wb = load_workbook(SUPPLIERS_FILE)
        ws = wb.active
        suppliers = []
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            data = dict(zip(headers, row))
            supplier = Supplier(
                id=str(data.get("id", "")),
                name=str(data.get("name", "")),
                contact_person=str(data.get("contact_person", "") or ""),
                email=str(data.get("email", "") or ""),
                phone=str(data.get("phone", "") or ""),
                notes=str(data.get("notes", "") or ""),
                purchase_volume=safe_float(data.get("purchase_volume")),
                country=str(data.get("country", "") or ""),
            )
            suppliers.append(supplier)
        wb.close()
        return suppliers

    def save_suppliers(self, suppliers: List[Supplier]):
        wb = Workbook()
        ws = wb.active
        ws.append(SUPPLIER_COLUMNS)
        for s in suppliers:
            ws.append([s.id, s.name, s.contact_person, s.email, s.phone, s.notes, s.purchase_volume, s.country])
        wb.save(SUPPLIERS_FILE)

    def add_supplier(self, supplier: Supplier):
        suppliers = self.load_suppliers()
        suppliers.append(supplier)
        self.save_suppliers(suppliers)

    def update_supplier(self, supplier: Supplier):
        suppliers = self.load_suppliers()
        for i, s in enumerate(suppliers):
            if s.id == supplier.id:
                suppliers[i] = supplier
                break
        self.save_suppliers(suppliers)

    def delete_supplier(self, supplier_id: str):
        suppliers = [s for s in self.load_suppliers() if s.id != supplier_id]
        self.save_suppliers(suppliers)

    # --- Users ---

    def _ensure_users_file(self):
        """Create users.xlsx with all three sheets and defaults if it doesn't exist."""
        if os.path.exists(USERS_FILE):
            return
        wb = Workbook()
        # Sheet 1: Users
        ws1 = wb.active
        ws1.title = "Users"
        ws1.append(["username", "display_name", "department", "role", "active"])
        # Sheet 2: Modulzugriff
        ws2 = wb.create_sheet("Modulzugriff")
        ws2.append(["department"] + MODULES)
        for dept, mods in _DEFAULT_MODULE_ACCESS.items():
            ws2.append([dept] + [mods.get(m, False) for m in MODULES])
        # Sheet 3: Aktionsrechte
        ws3 = wb.create_sheet("Aktionsrechte")
        ws3.append(["role"] + ACTION_PERMISSIONS)
        for role, perms in _DEFAULT_ACTION_RIGHTS.items():
            ws3.append([role] + [perms.get(p, False) for p in ACTION_PERMISSIONS])
        wb.save(USERS_FILE)

    def _open_users_wb(self):
        self._ensure_users_file()
        return load_workbook(USERS_FILE)

    def load_users(self) -> list[dict]:
        wb = self._open_users_wb()
        ws = wb["Users"]
        headers = [cell.value for cell in ws[1]]
        users = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            data = dict(zip(headers, row))
            data["active"] = bool(data.get("active", True))
            users.append(data)
        wb.close()
        return users

    def save_users(self, users: list[dict]):
        wb = self._open_users_wb()
        ws = wb["Users"]
        # Clear existing data rows
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
        # Rewrite
        for i, u in enumerate(users, start=2):
            ws.cell(i, 1, u.get("username", ""))
            ws.cell(i, 2, u.get("display_name", ""))
            ws.cell(i, 3, u.get("department", ""))
            ws.cell(i, 4, u.get("role", ""))
            ws.cell(i, 5, bool(u.get("active", True)))
        wb.save(USERS_FILE)
        wb.close()

    def load_module_access(self) -> dict:
        """Returns {department: {module: bool}}"""
        wb = self._open_users_wb()
        ws = wb["Modulzugriff"]
        headers = [cell.value for cell in ws[1]]
        modules = headers[1:]
        result = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            dept = row[0]
            result[dept] = {modules[j]: bool(row[j + 1]) for j in range(len(modules))}
        wb.close()
        # Fill missing departments with defaults
        for dept in DEPARTMENTS:
            if dept not in result:
                result[dept] = dict(_DEFAULT_MODULE_ACCESS.get(dept, {m: False for m in MODULES}))
        return result

    def save_module_access(self, matrix: dict):
        """matrix: {department: {module: bool}}"""
        wb = self._open_users_wb()
        ws = wb["Modulzugriff"]
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
        for i, dept in enumerate(DEPARTMENTS, start=2):
            ws.cell(i, 1, dept)
            for j, mod in enumerate(MODULES, start=2):
                ws.cell(i, j, bool(matrix.get(dept, {}).get(mod, False)))
        wb.save(USERS_FILE)
        wb.close()

    def load_action_rights(self) -> dict:
        """Returns {role: {permission: bool}}"""
        wb = self._open_users_wb()
        ws = wb["Aktionsrechte"]
        headers = [cell.value for cell in ws[1]]
        perms = headers[1:]
        result = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            role = row[0]
            result[role] = {perms[j]: bool(row[j + 1]) for j in range(len(perms))}
        wb.close()
        for role in ROLES:
            if role not in result:
                result[role] = dict(_DEFAULT_ACTION_RIGHTS.get(role, {p: False for p in ACTION_PERMISSIONS}))
        return result

    def save_action_rights(self, matrix: dict):
        """matrix: {role: {permission: bool}}"""
        wb = self._open_users_wb()
        ws = wb["Aktionsrechte"]
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
        for i, role in enumerate(ROLES, start=2):
            ws.cell(i, 1, role)
            for j, perm in enumerate(ACTION_PERMISSIONS, start=2):
                ws.cell(i, j, bool(matrix.get(role, {}).get(perm, False)))
        wb.save(USERS_FILE)
        wb.close()
