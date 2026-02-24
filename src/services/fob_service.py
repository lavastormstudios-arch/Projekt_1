import configparser
import os
import uuid
from typing import List, Optional

from src.models.fob_entry import FobEntry
from src.utils.constants import FOB_KALKULATION_FILE

_CONFIG_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "config.ini",
)

_FOB_DEFAULTS = {
    "eur_usd": "1.16",
    "eur_rmb": "8.2443",
    "fracht_40hc": "4300.0",
    "fracht_20": "3200.0",
    "zinssatz_pa": "0.0368",
    "frachtzeit_tage": "45",
    "rekla_quote": "0.015",
}


class FobService:
    def __init__(self, store):
        self._store = store

    # ------------------------------------------------------------------
    # Config
    # ------------------------------------------------------------------

    def get_config(self) -> dict:
        cfg = configparser.ConfigParser()
        cfg.read(_CONFIG_PATH, encoding="utf-8")
        result = {}
        for key, default in _FOB_DEFAULTS.items():
            result[key] = cfg.get("FOB", key, fallback=default)
        return result

    def save_config(self, values: dict):
        cfg = configparser.ConfigParser()
        cfg.read(_CONFIG_PATH, encoding="utf-8")
        if not cfg.has_section("FOB"):
            cfg.add_section("FOB")
        for key, val in values.items():
            cfg.set("FOB", key, str(val))
        with open(_CONFIG_PATH, "w", encoding="utf-8") as fh:
            cfg.write(fh)

    def _cfg_float(self, cfg: dict, key: str) -> float:
        try:
            return float(cfg.get(key, _FOB_DEFAULTS[key]))
        except (ValueError, TypeError):
            return float(_FOB_DEFAULTS[key])

    def _cfg_int(self, cfg: dict, key: str) -> int:
        try:
            return int(cfg.get(key, _FOB_DEFAULTS[key]))
        except (ValueError, TypeError):
            return int(_FOB_DEFAULTS[key])

    # ------------------------------------------------------------------
    # Calculations
    # ------------------------------------------------------------------

    def calculate(self, entry: FobEntry) -> dict:
        cfg = self.get_config()
        eur_usd = self._cfg_float(cfg, "eur_usd")
        eur_rmb = self._cfg_float(cfg, "eur_rmb")
        fracht_40hc = self._cfg_float(cfg, "fracht_40hc")
        fracht_20 = self._cfg_float(cfg, "fracht_20")
        zinssatz_pa = self._cfg_float(cfg, "zinssatz_pa")
        frachtzeit_tage = self._cfg_int(cfg, "frachtzeit_tage")
        rekla_quote = self._cfg_float(cfg, "rekla_quote")

        zinssatz_pro_tag = zinssatz_pa / 365.0
        lcl_pro_m3 = (fracht_40hc * 1.32) / 70.0

        # Y: EK in €  (Euro direkt, kein Wechselkurs)
        if entry.ek_fob_euro > 0:
            ek_in_eur = entry.ek_fob_euro
        elif entry.ek_fob_dollar > 0:
            ek_in_eur = entry.ek_fob_dollar / eur_usd
        elif entry.ek_fob_rmb > 0:
            ek_in_eur = entry.ek_fob_rmb / eur_rmb
        else:
            ek_in_eur = 0.0

        # Z: Finanzierungskosten
        finanzierungskosten = (entry.produktionszeit + frachtzeit_tage) * zinssatz_pro_tag * ek_in_eur

        # AA: Frachtkosten
        if entry.container_20 > 0:
            frachtkosten = fracht_20 / entry.container_20
        elif entry.container_40hc > 0:
            frachtkosten = fracht_40hc / entry.container_40hc
        else:
            frachtkosten = 0.0

        # AB: Reklakosten
        reklakosten = ek_in_eur * rekla_quote

        # AC: Kubikkosten (LCL)
        if entry.lcl and entry.kubikmeter > 0:
            kubikkosten = lcl_pro_m3 * entry.kubikmeter
        else:
            kubikkosten = 0.0

        # AD: Zollkosten
        zollkosten = (ek_in_eur + frachtkosten) * entry.zollsatz

        # Sonder-/Toolingkosten pro Stück
        if entry.container_40hc > 0:
            sonder_pro_stueck = entry.sonder_toolingkosten / entry.container_40hc
        elif entry.container_20 > 0:
            sonder_pro_stueck = entry.sonder_toolingkosten / entry.container_20
        else:
            sonder_pro_stueck = 0.0

        # I: NEUER EK
        neuer_ek = (
            ek_in_eur
            + finanzierungskosten
            + kubikkosten
            + frachtkosten
            + reklakosten
            + zollkosten
            + sonder_pro_stueck
        )

        # M: Marge UVP
        if entry.geplanter_uvp > 0:
            uvp_netto = entry.geplanter_uvp / 1.19
            marge_uvp = (uvp_netto - neuer_ek) / uvp_netto if uvp_netto != 0 else 0.0
        else:
            marge_uvp = 0.0

        # O: Marge Aktionspreis
        if entry.aktionspreis > 0:
            aktion_netto = entry.aktionspreis / 1.19
            marge_aktion = (aktion_netto - neuer_ek) / aktion_netto if aktion_netto != 0 else 0.0
        else:
            marge_aktion = 0.0

        return {
            "ek_in_eur": ek_in_eur,
            "finanzierungskosten": finanzierungskosten,
            "frachtkosten": frachtkosten,
            "reklakosten": reklakosten,
            "kubikkosten": kubikkosten,
            "zollkosten": zollkosten,
            "neuer_ek": neuer_ek,
            "marge_uvp": marge_uvp,
            "marge_aktion": marge_aktion,
        }

    # ------------------------------------------------------------------
    # CRUD
    # ------------------------------------------------------------------

    def load_all(self, include_archiv: bool = False) -> List[FobEntry]:
        entries = self._store.load_fob_entries()
        if not include_archiv:
            entries = [e for e in entries if not e.archiv]
        return entries

    def get_by_id(self, entry_id: str) -> Optional[FobEntry]:
        for e in self._store.load_fob_entries():
            if e.id == entry_id:
                return e
        return None

    def add(self, entry: FobEntry):
        if not entry.id:
            entry.id = str(uuid.uuid4())[:8].upper()
        self._store.add_fob_entry(entry)

    def update(self, entry: FobEntry):
        self._store.update_fob_entry(entry)

    def delete(self, entry_id: str):
        self._store.delete_fob_entry(entry_id)

    # ------------------------------------------------------------------
    # SAP Import
    # ------------------------------------------------------------------

    def import_from_excel(self, path: str) -> dict:
        """
        Read a SAP-exported Excel file and import/update FobEntry records.
        Matches by artnr. Returns {"new": N, "updated": M, "skipped": K}.
        """
        from openpyxl import load_workbook
        from src.utils.date_helpers import safe_float, safe_int

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
        except Exception as exc:
            raise ValueError(f"Datei konnte nicht geöffnet werden: {exc}") from exc

        headers_raw = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]

        # Flexible column mapping (case-insensitive, partial match)
        _col_map = {
            "artnr":                  ["artnr", "art.-nr", "artikelnummer", "artikel-nr"],
            "bezeichnung":            ["bezeichnung", "bezeichng", "artbezeichnung"],
            "lieferant":              ["lieferant", "lief."],
            "warengruppe":            ["warengruppe", "wgr", "wg"],
            "cm":                     ["cm"],
            "aktuelle_ztn":           ["aktuelle ztn", "ztn"],
            "aktueller_ek":           ["aktueller ek", "ek", "einkaufspreis"],
            "geplanter_uvp":          ["geplanter uvp", "uvp"],
            "aktionspreis":           ["aktionspreis", "aktion"],
            "ek_fob_dollar":          ["ek fob dollar", "ek fob $", "fob dollar", "fob $"],
            "ek_fob_rmb":             ["ek fob rmb", "fob rmb", "ek fob ¥"],
            "produktionszeit":        ["produktionszeit", "prod.zeit", "produktionszeit in tage"],
            "kubikmeter":             ["kubikmeter", "m³", "kubik"],
            "lcl":                    ["lcl"],
            "container_20":           ['20"', "20 zoll", "container 20", "20\""],
            "container_40hc":         ['40"hc', "40hc", "container 40hc", '40"', "40 hc"],
            "zollsatz":               ["zollsatz", "zoll"],
            "sonder_toolingkosten":   ["sonder", "tooling", "sonder/toolingkosten"],
            "archiv":                 ["archiv"],
        }

        def find_col(field_key: str) -> Optional[int]:
            synonyms = _col_map.get(field_key, [])
            for idx, h in enumerate(headers_raw):
                hl = h.lower()
                for syn in synonyms:
                    if syn.lower() in hl or hl in syn.lower():
                        return idx
            return None

        col_indices = {field: find_col(field) for field in _col_map}
        artnr_col = col_indices.get("artnr")
        if artnr_col is None:
            raise ValueError("Spalte 'Artnr' wurde nicht gefunden. Import abgebrochen.")

        existing_entries = self._store.load_fob_entries()
        existing_by_artnr = {e.artnr.strip().upper(): e for e in existing_entries}

        new_count = 0
        updated_count = 0
        skipped_count = 0

        def cell_val(row, idx):
            if idx is None:
                return None
            try:
                return row[idx].value
            except IndexError:
                return None

        for row in ws.iter_rows(min_row=2):
            raw_artnr = cell_val(row, artnr_col)
            if raw_artnr is None:
                continue
            artnr = str(raw_artnr).strip()
            if not artnr:
                continue

            key = artnr.upper()
            is_update = key in existing_by_artnr

            if is_update:
                e = existing_by_artnr[key]
            else:
                e = FobEntry(
                    id=str(uuid.uuid4())[:8].upper(),
                    artnr=artnr,
                    bezeichnung="",
                    lieferant="",
                )

            changed = False

            def assign_str(field, default=""):
                nonlocal changed
                idx = col_indices.get(field)
                val = cell_val(row, idx)
                if val is not None:
                    v = str(val).strip()
                    if getattr(e, field, default) != v:
                        setattr(e, field, v)
                        changed = True

            def assign_float(field, default=0.0):
                nonlocal changed
                idx = col_indices.get(field)
                val = cell_val(row, idx)
                if val is not None:
                    v = safe_float(val)
                    if getattr(e, field, default) != v:
                        setattr(e, field, v)
                        changed = True

            def assign_int(field, default=0):
                nonlocal changed
                idx = col_indices.get(field)
                val = cell_val(row, idx)
                if val is not None:
                    v = safe_int(val)
                    if getattr(e, field, default) != v:
                        setattr(e, field, v)
                        changed = True

            def assign_bool(field, default=False):
                nonlocal changed
                idx = col_indices.get(field)
                val = cell_val(row, idx)
                if val is not None:
                    v = bool(val)
                    if getattr(e, field, default) != v:
                        setattr(e, field, v)
                        changed = True

            assign_str("bezeichnung")
            assign_str("lieferant")
            assign_str("warengruppe")
            assign_str("cm")
            assign_str("aktuelle_ztn")
            assign_float("aktueller_ek")
            assign_float("geplanter_uvp")
            assign_float("aktionspreis")
            assign_float("ek_fob_dollar")
            assign_float("ek_fob_rmb")
            assign_int("produktionszeit")
            assign_float("kubikmeter")
            assign_bool("lcl")
            assign_int("container_20")
            assign_int("container_40hc")
            assign_float("zollsatz")
            assign_float("sonder_toolingkosten")
            assign_bool("archiv")

            if is_update:
                if changed:
                    self._store.update_fob_entry(e)
                    updated_count += 1
                else:
                    skipped_count += 1
            else:
                self._store.add_fob_entry(e)
                new_count += 1

        wb.close()
        return {"new": new_count, "updated": updated_count, "skipped": skipped_count}
